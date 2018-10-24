# python scrip for tracker update.
# 1. To update statistic data in the tracker for daily/weekly report purpose
# 2. To prepare and wash the data for Power BI purpose


# version 1.0
# 2018-10-17

import sys


from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

# 1. To update statistic data in the tracker for daily/weekly report purpose
#read data from the tracker, read the sites in scope only

# 尝试了三种方法把处理过的数据导入Pandas Dataframe 或者 Python Dictionary
# a. 用openpyxl打开，通过条件判断过滤掉不需要的数据，并同时存储到相应的dict中。处理过程相对较慢，字典的key是以行为单位的，每行的Site ID。
# b. 用Pandas的read_excel打开，直接读取所有的内容，然后转换成字典格式（.to_dict()）， 然后用pop() 把不需要的行剔除出去。这时的dict是以列为单位的。
# c. 用Pandas的read_excel打开并在打开时，使用df.loc选择需要的行列，直接生成dataframe。
# 最终选择了第三种，df.loc的使用大大提升了效率和逻辑清晰度。
# 最后输出过滤过的原始数据dataframe
def readTracker(filename):
		
	import pandas as pd
	
	df = pd.read_excel(filename, 'SC Site Level Info')
	tracker = df.loc[df['Version 1.5'] == 'Version 1.5']
	
	''' get the tracker using Dictionary ,old school :(
	allData = pd.read_excel(filename,'SC Site Level Info', index_col = 0).to_dict()
	originalData = allData
	
	# remove on hold sites 
	onHoldList = [key for key,val in originalData['Version 1.5'].items() if val != 'Version 1.5']
	for item in onHoldList:
		for key,val in originalData.items():
			
			originalData[key].pop(item)
	'''
	
	return tracker;

# udpate statistic in tracker, including daily report and weekly report.
# 根据条件和输入的基础数据，进行统计。
# 这里维护了一张统计时所需要依据的条件的表格。表格内包括：原始数据的表头与统计数据表头的对应关系，以及是以何种条件进行筛选
# 此段代码主要通过df.loc, len(), pd.notnull() 进行统计
# 最后输出一张包括含统计数据的dataframe
def getStatistic(filename,mappingFile,tracker):
	
	import pandas as pd
	import openpyxl
	
	mappingTable = pd.read_excel(mappingFile,'Sheet1')
	
	mappingTable.loc[mappingTable['Column Name in Tracker'] == 0, ['Number of Sites']] = 0
	
	for item in mappingTable.loc[mappingTable['Condition'] == 'Not Null', 'Name in Report']:
		nameInTracker = mappingTable.loc[mappingTable['Name in Report'] == item, 'Column Name in Tracker'].tolist()[0]
		mappingTable.loc[mappingTable['Name in Report'] == item,'Number of Sites'] = len(tracker[pd.notnull(tracker[nameInTracker])])
		
	for item in mappingTable.loc[mappingTable['Condition'] == 'Match', 'Name in Report']:
		nameInTracker = mappingTable.loc[mappingTable['Name in Report'] == item, 'Column Name in Tracker'].tolist()[0]
		matchValue = mappingTable.loc[mappingTable['Name in Report'] == item, 'Match Value'].tolist()[0]
		mappingTable.loc[mappingTable['Name in Report'] == item,'Number of Sites'] = len(tracker[tracker[nameInTracker] == matchValue])
		
	#to align with current tracker format, FORM A / EME Passed sites number should be passed by Cluster HLD
	mappingTable.loc[mappingTable['Name in Report'] == 'Form A','Number of Sites'] = len(tracker[(tracker['Form A Submitted'] == 'Y') & (pd.notnull(tracker['RF Lock Down']))])
	mappingTable.loc[mappingTable['Name in Report'] == 'EME Passed','Number of Sites'] = len(tracker[(pd.notnull(tracker['Form A Final Status'])) & (pd.notnull(tracker['RF Lock Down']))])
	
	
	return mappingTable;

# 将统计好的数据，填入到表格之后。主要使用openpyxl。缺点是速度有些慢，需要一个一个的填。目前还没有想到好的解决办法
# 最后存储填好的表格。
def updateTracker(filename, mappingTable,mappingFile):
	
	import openpyxl
	import time
	import pandas as pd
	
	wb = openpyxl.load_workbook(filename = filename)
	
	#Update Form A Tracker
	ws = wb['Form A Tracker']
	nextRow = ws.max_row + 1
	
	for column in range (2, ws.max_column + 1):
		currentTitle = ws.cell(row = 2, column = column).value
		ws.cell(row = nextRow, column = column).value = mappingTable.loc[mappingTable['Name in Report'] == currentTitle, 'Number of Sites'].tolist()[0]
		
	ws.cell(row = nextRow, column = 1).value = time.strftime("%d/%m/%Y")
	
	#Update Deployment tracker
	ws = wb['Small Cell Deployment tracker']
	nextRow = ws.max_row + 1
	
	for column in range (2, 11):
		currentTitle = ws.cell(row = 3, column = column).value
		ws.cell(row = nextRow, column = column).value = mappingTable.loc[mappingTable['Name in Report'] == currentTitle, 'Number of Sites'].tolist()[0]
		
	ws.cell(row = nextRow, column = 1).value = time.strftime("%d/%m/%Y")
	
	
	#update weekly tracker
	df = pd.read_excel (filename, 'SC Site Level Info')
	
	phase1 = df.loc[df['Rollout Priority'] == 'Phase 1']
	phase2 = df.loc[df['Rollout Priority'] == 'Phase 2']
	phase3 = df.loc[df['Rollout Priority'] == 'Phase 3']
	
	
	phase1Value = getStatistic(filename,mappingFile,phase1)
	phase2Value = getStatistic(filename,mappingFile,phase2)
	phase3Value = getStatistic(filename,mappingFile,phase3)
	
	ws = wb['Cluster Stats - Phase']
	
	for column in range (22,33):
		currentTitle = ws.cell (row = 1, column = column).value
		ws.cell(row = 3, column = column).value = phase1Value.loc[phase1Value['Name in Report'] == currentTitle, 'Number of Sites'].tolist()[0]
		ws.cell(row = 4, column = column).value = phase2Value.loc[phase1Value['Name in Report'] == currentTitle, 'Number of Sites'].tolist()[0]
		ws.cell(row = 5, column = column).value = phase3Value.loc[phase1Value['Name in Report'] == currentTitle, 'Number of Sites'].tolist()[0]
	
	ws.cell(row = 2, column = 2).value = time.strftime("%d/%m/%Y")
		
	wb.save(filename)

# 第一段代码的主函数，通过键入日期来确定打开的文件名
def trackerUpdate ():
	
	filename = 'Master Site List '+ foo.form_widget.dateTypeIn.text()+'.xlsx'
	
	mappingFile = 'columnMapping.xlsx'
	tracker = readTracker(filename)
	mappingTable = getStatistic(filename, mappingFile, tracker)
	updateTracker(filename,mappingTable,mappingFile)


# 2. To prepare and wash the data for Power BI purpose

# wash the list
# 此处分为三步，1）将原始表格进行再一步的过滤，因为在power BI内不需要所有的列，只保留需要的列。 
#              2）根据规则，增加新的一列，统计出各个站点当前所处的阶段。使用了df.apply来实现，apply后边的加的函数，以列为单位
#              3）由于需要做柱状图，需要统一legend， 将之前不同milestone不同的说明归一化。使用df.loc 进行dataframe的重新赋值

def washTracker(tracker,mappingFile):
	
	import pandas as pd
	
	validColumn = pd.read_excel(mappingFile,'Sheet2')
	cols = [c for c in tracker.columns if c in validColumn.values]
	trackerBI = tracker[cols]
	
	trackerBI['Site Status'] = trackerBI.apply(siteStatus, axis = 1)
	trackerBI['Site Status Power BI'] = trackerBI.apply(siteStatusPowerBI, axis = 1)
	
	milestone = pd.read_excel(mappingFile,'Sheet3')
	milestoneColumn = milestone['Column in Report'].tolist()
	
	for item in milestoneColumn:
		trackerBI.loc[pd.notnull(tracker[item]), item] = 'Done'
		trackerBI.loc[pd.isnull(tracker[item]), item] = 'To be completed'
		
	
	return trackerBI;

# 配合df.apply使用的函数
def siteStatusPowerBI(row):
	
	import pandas as pd
	if pd.notnull(row['Commissioning & Integration']): return '6 - Commissioning and Integration'
	if pd.notnull(row['RFI Status']): return '5 - RFI Report'
	if pd.notnull(row['Site Installed']) : return '4 - Equipment Install Complete'
	if pd.notnull(row['RFNSA STAD table locked']): return '3 - STAD Table Locked'
	if pd.notnull(row['RF Lock Down']): return '2 - Cluster Finalization'
	return '1 - MSL Released'

# 配合df.apply使用的函数（主要用于跟新每周tracker的kml）
def siteStatus(row):
	
	import pandas as pd
	if pd.notnull(row['Commissioning & Integration']): return 'Commissioning and Integration'
	if pd.notnull(row['RFI Status']): return 'RFI Report'
	if pd.notnull(row['Site Installed']) : return 'Equipment Install Complete'
	if pd.notnull(row['RFNSA STAD table locked']): return 'STAD Table Locked'
	if pd.notnull(row['RF Lock Down']): return 'Cluster Finalization'
	return 'MSL Released'
	
# 第二段函数的主函数
# 使用pd.ExcelWriter 将dataframe写入一个给定位置的excel。每天更新的时候回不断覆盖之前的内容
# writer = pd.ExcelWriter(out_path, engine = 'xlsxwriter'), df.to_excel(writer,sheet_name = xxx, index = False), writer.save() 
def createPowerBIFile():
	
	import xlsxwriter, pandas as pd
	
	filename = 'Master Site List '+ foo.form_widget.dateTypeIn.text()+'.xlsx'
	mappingFile = 'columnMapping.xlsx'
	tracker = readTracker(filename)
	
	trackerBI = washTracker(tracker,mappingFile)
	out_path = r"C:\Caspar Doc\000 Master Site List\01 SC & Macro Tracker\Tracker Report\PowerBIData.xlsx"
	writer = pd.ExcelWriter(out_path, engine = 'xlsxwriter')
	trackerBI.to_excel(writer,sheet_name ='Power BI Data',index = False)
	writer.save()

	
class MyMainWindow(QMainWindow):
		
		def __init__(self, parent = None):
				
				super(MyMainWindow, self).__init__(parent)
				self.form_widget = FormWidget(self)
				self.setCentralWidget(self.form_widget)
				
							
				#set Exit Button
				exitAct = QAction(QIcon('exit.png'), 'Exit', self)
				exitAct.setShortcut('Ctrl+Q')
				exitAct.setStatusTip('Exit application')
				exitAct.triggered.connect(self.close)
				
				#show StatusBar
				self.statusBar()
				
				#create MenuBar
				menubar = self.menuBar()
				fileMenu = menubar.addMenu('&File')
				toolDes = QAction('This is a tracker updating tool',self)
				fileMenu.addAction(toolDes)
				
				fileMenu.addAction(exitAct)
				
				#create ToolBar
				toolbar = self.addToolBar('Exit')
				toolbar.addAction(exitAct)
				
				#set popped up window location/title
				self.setGeometry(300,300,250,250)
				self.setWindowTitle('Daily Tracker Update')
		
class FormWidget(QWidget):
		
		def __init__(self,parent):
				
				super(FormWidget,self).__init__(parent)
				
				self.layout = QVBoxLayout(self)
				
				
				#Date to type in
				self.dateTypeIn = QLineEdit()
				self.layout.addWidget(self.dateTypeIn)
								
				# Daily Update
				self.button1 = QPushButton("Daily update")
				self.layout.addWidget(self.button1)
				
				self.button1.clicked.connect(trackerUpdate)
				self.button1.clicked.connect(self.clickMethod)
				
				
				# create Power BI data and change it to excel format
				self.button2 = QPushButton("Create Power BI data")
				self.layout.addWidget(self.button2)
				
				self.button2.clicked.connect(createPowerBIFile)
				self.button2.clicked.connect(self.clickMethod)
				
				'''
				# Weekly Phase1/2/3 update
				self.button3 = QPushButton("Weekly Phase1/2/3 update")
				self.layout.addWidget(self.button3)
				
				self.button3.clicked.connect(f03_phaseUpdate)
				self.button3.clicked.connect(self.clickMethod)
				'''
				self.setLayout(self.layout)
				
		def clickMethod(self):
				QMessageBox.about(self,"Result","Done")

if __name__ == '__main__' :

		app = QApplication(sys.argv)
		foo = MyMainWindow()
		foo.show()
		sys.exit(app.exec_())
